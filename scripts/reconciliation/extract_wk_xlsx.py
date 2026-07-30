"""
One-off extraction of the real-world wk.xlsx workbook into JSON fixtures
consumed by tests/Fixtures/RashidWorkbook/*.json.

Not part of the app/CI; run manually whenever the workbook facts are revisited:

    python scripts/reconciliation/extract_wk_xlsx.py

Fails loudly (raises) on anything that doesn't match the facts this dataset
was already verified to contain, rather than silently miscoding data.
"""
import json
import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = REPO_ROOT / "wk.xlsx"
FIXTURES_DIR = REPO_ROOT / "tests" / "Fixtures" / "RashidWorkbook"

KNOWN_PROJECT_NAMES = {
    "تكية اطعام", "سقيا ماء", "رغيف خيري", "توزيع نقدي", "كسوة أطفال وكبار",
    "سلاس خضروات", "مواد تموينية", "طرود نظافة صحية", "بطاطين وحرامات", "خيام",
    "شوادر", "فرشات", "انشاء مرافق صحية", "أحذية", "حليب أطفال", "بمبرز",
    "إفطار المدرسة", "توزيع دقيق", "توزيع لحوم", "توزيع معجنات", "أدوات مطبخية",
    "معدات النظافة الشخصية", "حقائب عناية نسائية", "حمام", "توزيع حطب",
    "سولار لضخ المياه",
}


def to_date_string(value) -> str:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    raise ValueError(f"expected a date, got {value!r}")


def extract_daily_income_expense(wb) -> dict:
    """Day-sheets '1'..'30': rows 4-29, columns A(name)/C(income)/D(expense)."""
    result = {}
    for day in range(1, 31):
        ws = wb[str(day)]
        journal_date = to_date_string(ws["D2"].value)
        per_project = {}
        for row in range(4, 30):
            name = ws.cell(row=row, column=1).value
            if name is None:
                continue
            if name not in KNOWN_PROJECT_NAMES:
                raise ValueError(f"day {day} row {row}: unknown project name {name!r}")
            income = ws.cell(row=row, column=3).value or 0
            expense = ws.cell(row=row, column=4).value or 0
            per_project[name] = {"income": float(income), "expense": float(expense)}
        missing = KNOWN_PROJECT_NAMES - per_project.keys()
        if missing:
            raise ValueError(f"day {day}: missing projects {missing}")
        result[journal_date] = per_project
    return result


def extract_inventory_opening_batches(wb) -> list:
    """المخزون sheet, first section ('مواد خام (تشغيلي مشاريع)'), rows 6-75.

    Each item's own row carries its own introduction date in column A (not
    uniformly 2026-03-31 - a few items enter the catalog mid-month), so the
    per-row date is preserved as this item's opening-batch date.
    """
    ws = wb["المخزون"]
    items = []
    seen_codes = set()
    for row in range(6, 76):
        code = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value
        if code is None or name is None:
            continue
        opening_date = ws.cell(row=row, column=1).value
        quantity = ws.cell(row=row, column=4).value or 0
        unit = ws.cell(row=row, column=5).value
        unit_price = ws.cell(row=row, column=7).value or 0
        if code in seen_codes:
            raise ValueError(f"duplicate item code {code!r} at row {row}")
        seen_codes.add(code)
        items.append({
            "code": code,
            "name": name,
            "unit": unit,
            "opening_date": to_date_string(opening_date),
            "opening_quantity": float(quantity),
            "opening_unit_price": float(unit_price),
        })
    return items


EXPENSE_TYPE_MAP = {"تشغيلي": "operational", "إداري": "administrative"}


def extract_inventory_movements(wb) -> list:
    """حركات المخزون sheet, header row 10, data from row 11.

    Row order is preserved exactly as it appears in the sheet - this becomes
    the FIFO tie-break (insertion) order when multiple rows share a date.
    """
    ws = wb["حركات المخزون"]
    movements = []
    for row in range(11, ws.max_row + 1):
        movement_type = ws.cell(row=row, column=7).value  # G
        if movement_type not in ("وارد", "صادر"):
            continue
        movement_date = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        entry = {
            "date": to_date_string(movement_date),
            "item_code": code,
        }
        if movement_type == "وارد":
            quantity = ws.cell(row=row, column=9).value  # I
            unit_price = ws.cell(row=row, column=10).value  # J
            entry.update({
                "type": "incoming",
                "quantity": float(quantity),
                "unit_price": float(unit_price),
            })
        else:
            quantity = ws.cell(row=row, column=14).value  # N
            beneficiary = ws.cell(row=row, column=6).value  # F
            expense_type_raw = ws.cell(row=row, column=12).value  # L
            if expense_type_raw != "تشغيلي":
                raise ValueError(
                    f"row {row}: expected outgoing expense_type 'تشغيلي', "
                    f"got {expense_type_raw!r} - this dataset was verified to have "
                    f"NO administrative-typed outgoing movements; the "
                    f"'administrative_expense always 0' scope assumption depends on this"
                )
            if beneficiary is not None and beneficiary not in KNOWN_PROJECT_NAMES:
                raise ValueError(f"row {row}: unknown beneficiary project {beneficiary!r}")
            # ~9 of 201 outgoing rows leave the beneficiary column blank in the
            # source sheet (anonymous consumption). beneficiary_project=None
            # here means "attribute to the synthetic inventory-owner project" -
            # inert for every calculation under test since these rows are all
            # 'تشغيلي' (operational), never 'إداري' (administrative).
            entry.update({
                "type": "outgoing",
                "quantity": float(quantity),
                "beneficiary_project": beneficiary,
                "expense_type": EXPENSE_TYPE_MAP[expense_type_raw],
            })
        movements.append(entry)
    return movements


def remove_opening_balance_duplicates(opening_batches: list, movements: list) -> list:
    """المخزون's opening-balance row and حركات المخزون's same-day incoming row
    are, for most items, the SAME real-world stocking event recorded in two
    different sheets (identical date/quantity/unit_price) - not a separate
    second delivery. Keeping both would silently double an item's opening
    stock. Drop at most one exact-match incoming row per opening batch.
    """
    filtered = []
    for m in movements:
        if m["type"] != "incoming":
            filtered.append(m)
            continue

        is_duplicate = any(
            o["code"] == m["item_code"]
            and o["opening_date"] == m["date"]
            and o["opening_quantity"] == m["quantity"]
            and o["opening_unit_price"] == m["unit_price"]
            for o in opening_batches
        )
        if not is_duplicate:
            filtered.append(m)

    return filtered


def main() -> None:
    wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)

    daily = extract_daily_income_expense(wb)
    opening_batches = extract_inventory_opening_batches(wb)
    movements = extract_inventory_movements(wb)
    movements = remove_opening_balance_duplicates(opening_batches, movements)

    incoming_count = sum(1 for m in movements if m["type"] == "incoming")
    outgoing_count = sum(1 for m in movements if m["type"] == "outgoing")
    if incoming_count != 34 or outgoing_count != 201:
        raise ValueError(
            f"expected 34 incoming (96 minus 62 opening-balance duplicates) / "
            f"201 outgoing movements, got {incoming_count} incoming / {outgoing_count} outgoing"
        )

    (FIXTURES_DIR / "daily_income_expense.json").write_text(
        json.dumps(daily, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (FIXTURES_DIR / "inventory_opening_batches.json").write_text(
        json.dumps(opening_batches, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (FIXTURES_DIR / "inventory_movements.json").write_text(
        json.dumps(movements, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    print(f"daily_income_expense: {len(daily)} days x {len(KNOWN_PROJECT_NAMES)} projects")
    print(f"inventory_opening_batches: {len(opening_batches)} items")
    print(f"inventory_movements: {incoming_count} incoming, {outgoing_count} outgoing")


if __name__ == "__main__":
    main()
